"""
program that creates an excel spreadsheet with data from lists
"""

from openpyxl import Workbook

# example lists of data
favorite_foods = ['Pizza', 'Chocolate Cake', 'Broccoli']
favorite_colors = ['Blue', 'Purple', 'Green', 'Orange']

workbook = Workbook()  # assigning variable for the whole sheet so it's easier to use

worksheet = workbook.active  # assigning variable for the first page

worksheet.title = 'Favorite Things'  # naming the first page

worksheet.cell(1, 1, 'Favorite Foods')  # making the first specific cell say favorite foods
for index, food in enumerate(favorite_foods):  # loop that pastes list into the excel spreadsheet.
    worksheet.cell(index+2, 1, food)

worksheet.cell(1, 2, 'Favorite Colors')  # creating another column for colors list
for index, color in enumerate(favorite_colors):
    worksheet.cell(index+2, 2, color)

workbook.save('favorites.xlsx')  # saving the workbook once finished changing it
