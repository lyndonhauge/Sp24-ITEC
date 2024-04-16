"""
program that creates an excel spreadsheet with data from dictionaries
"""

from openpyxl import Workbook

# example dictionary, temperatures for a week
week_temperatures = {
    'Monday': 54,
    'Tuesday': 60,
    'Wednesday': 62,
    'Thursday': 57,
    'Friday': 71,
}

workbook = Workbook()

worksheet = workbook.active

worksheet.title = 'Daily Temperatures'

row_index = 2  # variable for row to put each cell in (where the loop will start writing)

worksheet.cell(1, 1, 'Day')
worksheet.cell(1, 2, 'Temperature (F)')
for day, temp in week_temperatures.items():  # looping over dictionary and writing it all into spreadsheet
    worksheet.cell(row_index, 1, day)
    worksheet.cell(row_index, 2, temp)
    row_index += 1  # quicker way of writing row_index = row_index + 1

workbook.save('temperatures.xlsx')
